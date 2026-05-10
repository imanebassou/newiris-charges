import json
import requests
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions

from .extractors.pdf_extractor import extract_from_pdf
from .extractors.image_extractor import extract_from_image
from .extractors.text_cleaner import clean_text
from .analyzers.ai_analyzer import analyze_document
from .validators.field_validator import validate_fields

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

SUPPORTED_FORMATS = {
    'pdf': 'pdf',
    'docx': 'word',
    'doc': 'word',
    'xlsx': 'excel',
    'xls': 'excel',
    'jpg': 'image',
    'jpeg': 'image',
    'png': 'image',
    'bmp': 'image',
    'tiff': 'image',
    'txt': 'text',
}


def extract_text(file) -> dict:
    filename = file.name.lower()
    ext = filename.rsplit('.', 1)[-1] if '.' in filename else ''
    file_format = SUPPORTED_FORMATS.get(ext)

    if not file_format:
        return {"text": "", "success": False, "error": f"Format .{ext} non supporté"}

    if file_format == 'pdf':
        result = extract_from_pdf(file)

    elif file_format == 'image':
        result = extract_from_image(file)

    elif file_format == 'word':
        try:
            import docx
            doc = docx.Document(file)
            text = '\n'.join([p.text for p in doc.paragraphs])
            result = {"text": clean_text(text), "success": True, "method": "python-docx"}
        except Exception as e:
            result = {"text": "", "success": False, "error": str(e)}

    elif file_format == 'excel':
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file)
            text = ""
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = ' | '.join([str(c) for c in row if c is not None])
                    if row_text.strip():
                        text += row_text + "\n"
            result = {"text": clean_text(text), "success": True, "method": "openpyxl"}
        except Exception as e:
            result = {"text": "", "success": False, "error": str(e)}

    elif file_format == 'text':
        try:
            file.seek(0)
            text = file.read().decode('utf-8', errors='ignore')
            result = {"text": clean_text(text), "success": True, "method": "text"}
        except Exception as e:
            result = {"text": "", "success": False, "error": str(e)}

    else:
        result = {"text": "", "success": False, "error": "Format non supporté"}

    return result


class DocumentExtractView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Aucun fichier fourni'}, status=400)

        if file.size > MAX_FILE_SIZE:
            return Response({'error': 'Fichier trop grand (max 10MB)'}, status=400)

        # 1. Extraction texte
        extraction = extract_text(file)
        if not extraction.get('success') or not extraction.get('text'):
            return Response({
                'error': extraction.get('error', 'Impossible d\'extraire le texte'),
                'filename': file.name,
                'analyse': {
                    "type_document": "autre",
                    "module_suggere": "autre",
                    "confiance": "faible",
                    "champs": {},
                    "resume": "Extraction échouée. Remplissez manuellement.",
                    "actions_suggerees": []
                }
            }, status=200)

        text = extraction['text']

        # 2. Analyse IA
        analysis = analyze_document(text, file.name)

        # 3. Validation champs
        champs_valides = validate_fields(analysis.get('champs', {}))
        analysis['champs_valides'] = champs_valides

        return Response({
            'filename': file.name,
            'format': file.name.rsplit('.', 1)[-1].upper(),
            'text_extrait': text[:600],
            'methode_extraction': extraction.get('method', 'unknown'),
            'analyse': analysis,
        })